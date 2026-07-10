"""
Generate Meesho bulk catalog upload sheet for Whiff Theory fragrances.

Run: python scripts/generate-meesho-catalog.py
Output: exports/meesho-bulk-catalog-whiff-theory.xlsx

Note: Meesho's official template columns vary by sub-category. After downloading
your category template from Supplier Panel, copy values from the Catalog Data
sheet if column names differ slightly.
"""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "exports" / "meesho-bulk-catalog-whiff-theory.xlsx"

MANDATORY_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red
RECOMMENDED_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
HEADER_FONT = Font(bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

BRAND = "Whiff Theory"
ORIGIN = "India"
GST = 18
HSN = "3303"  # Perfumes and toilet waters
MANUFACTURER = (
    "Whiff Theory, Visakhapatnam, Andhra Pradesh, India — "
    "Contact: WhiffTheoryStore@gmail.com"
)
PACKER = MANUFACTURER
CUSTOMER_CARE = "WhiffTheoryStore@gmail.com | whiff-theory.com"
IMAGE_PLACEHOLDER = "PASTE LINK FROM MEESHO IMAGE BULK UPLOADER (JPEG, RGB only)"

FRAGRANCES = [
    {
        "group": "WT-PASSION",
        "style": "WT-PASSION",
        "name": "Passion",
        "variation_20": "20 ml",
        "variation_50": "50 ml",
        "price_20": 300,
        "price_50": 799,
        "mrp_20": 399,
        "mrp_50": 999,
        "weight_20": 95,
        "weight_50": 165,
        "tags": "Oriental · Fruity",
        "short": "Warm saffron and passionfruit with oud, amber and vanilla.",
        "ideal_for": "Unisex",
    },
    {
        "group": "WT-IRISH-TWEED",
        "style": "WT-IRISH-TWEED",
        "name": "Irish Tweed",
        "variation_20": "20 ml",
        "variation_50": "50 ml",
        "price_20": 300,
        "price_50": 799,
        "mrp_20": 399,
        "mrp_50": 999,
        "weight_20": 95,
        "weight_50": 165,
        "tags": "Woody · Green · Aquatic",
        "short": "Crisp verbena, iris and sandalwood for day to evening wear.",
        "ideal_for": "Unisex",
    },
    {
        "group": "WT-PETALINA",
        "style": "WT-PETALINA",
        "name": "Petalina",
        "variation_20": "20 ml",
        "variation_50": "50 ml",
        "price_20": 249,
        "price_50": 599,
        "mrp_20": 349,
        "mrp_50": 749,
        "weight_20": 95,
        "weight_50": 165,
        "tags": "Floral · Fruity · Musk",
        "short": "Juicy lychee and Turkish rose with white musk and soft woods.",
        "ideal_for": "Women",
    },
    {
        "group": "WT-APHRODITE",
        "style": "WT-APHRODITE",
        "name": "Aphrodite",
        "variation_20": "20 ml",
        "variation_50": "50 ml",
        "price_20": 399,
        "price_50": 1099,
        "mrp_20": 499,
        "mrp_50": 1299,
        "weight_20": 95,
        "weight_50": 165,
        "tags": "Citrus · Fruity · Oriental",
        "short": "Sicilian citrus and white florals into amber warmth.",
        "ideal_for": "Women",
    },
    {
        "group": "WT-SCORCHED",
        "style": "WT-SCORCHED",
        "name": "Scorched",
        "variation_20": "20 ml",
        "variation_50": "50 ml",
        "price_20": 449,
        "price_50": 1199,
        "mrp_20": 549,
        "mrp_50": 1399,
        "weight_20": 95,
        "weight_50": 165,
        "tags": "Smoky · Leather · Spicy",
        "short": "Birch tar, leather and incense. Bold evening scent.",
        "ideal_for": "Unisex",
    },
    {
        "group": "WT-FERAL",
        "style": "WT-FERAL",
        "name": "Feral",
        "variation_20": "20 ml",
        "variation_50": "50 ml",
        "price_20": 499,
        "price_50": 1299,
        "mrp_20": 599,
        "mrp_50": 1499,
        "weight_20": 95,
        "weight_50": 165,
        "tags": "Animalic · Woody · Spicy",
        "short": "Cumin, oud and cedar. Polarising extrait for fabric.",
        "ideal_for": "Unisex",
    },
]

# Column order aligned with common Meesho beauty/fragrance bulk templates
COLUMNS = [
    ("Product Name", "mandatory"),
    ("Variation", "mandatory"),
    ("Meesho Price", "mandatory"),
    ("Wrong/Defective Returns Price", "mandatory"),
    ("MRP", "mandatory"),
    ("GST %", "mandatory"),
    ("HSN ID", "mandatory"),
    ("Product weight (gms)", "mandatory"),
    ("Inventory", "mandatory"),
    ("Country of Origin", "mandatory"),
    ("Manufacturer Details", "mandatory"),
    ("Packer Details", "mandatory"),
    ("Ideal For", "recommended"),
    ("Fragrance Type", "recommended"),
    ("Multipack", "recommended"),
    ("Image 1 (Front)", "mandatory"),
    ("Product ID / Style ID", "recommended"),
    ("SKU ID", "recommended"),
    ("Image 2", "recommended"),
    ("Image 3", "recommended"),
    ("Image 4", "recommended"),
    ("Brand Name", "recommended"),
    ("Group ID", "recommended"),
    ("Product Description", "recommended"),
    ("Capacity", "recommended"),
    ("Net Quantity", "recommended"),
    ("Concentration", "recommended"),
    ("Application", "recommended"),
    ("Customer Care", "recommended"),
    ("Importer Details", "optional"),
]


def returns_price(meesho_price: int) -> int:
    return max(1, math.floor(meesho_price * 0.9))


def product_title(name: str, ml: str) -> str:
    return f"Whiff Theory {name} Apparel Perfume {ml}"


def description(frag: dict, ml: str) -> str:
    return (
        f"{frag['short']} 25% extrait. Type: apparel perfume — spray on clothing, "
        f"not skin. Net qty: {ml}. Notes: {frag['tags']}. Crafted in Vizag, India. "
        f"MRP incl. taxes. {CUSTOMER_CARE}"
    )


def build_rows() -> list[dict]:
    rows: list[dict] = []
    for frag in FRAGRANCES:
        for ml_key, price_key, mrp_key, weight_key, var_key in [
            ("20 ml", "price_20", "mrp_20", "weight_20", "variation_20"),
            ("50 ml", "price_50", "mrp_50", "weight_50", "variation_50"),
        ]:
            meesho = frag[price_key]
            rows.append(
                {
                    "Product Name": product_title(frag["name"], ml_key),
                    "Variation": frag[var_key],
                    "Meesho Price": meesho,
                    "Wrong/Defective Returns Price": returns_price(meesho),
                    "MRP": frag[mrp_key],
                    "GST %": GST,
                    "HSN ID": HSN,
                    "Product weight (gms)": frag[weight_key],
                    "Inventory": 25,
                    "Country of Origin": ORIGIN,
                    "Manufacturer Details": MANUFACTURER,
                    "Packer Details": PACKER,
                    "Ideal For": frag["ideal_for"],
                    "Fragrance Type": "Eau de Parfum",
                    "Multipack": 1,
                    "Image 1 (Front)": IMAGE_PLACEHOLDER,
                    "Product ID / Style ID": frag["style"],
                    "SKU ID": f"{frag['style']}-{ml_key.replace(' ', '').upper()}",
                    "Image 2": IMAGE_PLACEHOLDER,
                    "Image 3": "",
                    "Image 4": "",
                    "Brand Name": BRAND,
                    "Group ID": frag["group"],
                    "Product Description": description(frag, ml_key),
                    "Capacity": ml_key,
                    "Net Quantity": ml_key,
                    "Concentration": "25%",
                    "Application": "Spray on clothing — shoulders, arms, torso",
                    "Customer Care": CUSTOMER_CARE,
                    "Importer Details": "",
                }
            )
    return rows


def style_header(ws, row: int = 1) -> None:
    for col_idx, (name, level) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = MANDATORY_FILL if level == "mandatory" else (
            RECOMMENDED_FILL if level == "recommended" else PatternFill()
        )
        cell.alignment = WRAP


def autosize(ws, max_width: int = 42) -> None:
    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max_width, max(12, len(name) + 2))


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Instructions
    inst = wb.active
    inst.title = "Instructions"
    instructions = [
        "Whiff Theory — Meesho Bulk Catalog Upload",
        "",
        "BEFORE UPLOADING",
        "1. Download the official Meesho template for your exact sub-category",
        "   (e.g. Beauty & Personal Care → Fragrance → Perfume) from Supplier Panel.",
        "2. Copy rows from 'Catalog Data' into that template if column names differ.",
        "3. Upload product images via Supplier Panel → Images Bulk Upload.",
        "   Paste those links into Image 1–4 columns (no Google Drive links).",
        "4. Select Variation / Capacity / HSN from dropdowns in Meesho template only.",
        "",
        "PRICING RULES",
        "• Meesho Price and Wrong/Defective Returns Price must both be < MRP",
        "• Wrong/Defective Returns Price must be < Meesho Price",
        "",
        "IMAGE RULES",
        "• Min 1 front image per SKU; packaging front + back recommended",
        "• JPEG only, RGB colour space, no watermark/text on primary image",
        "",
        "VARIANTS",
        "• One row per size (20 ml and 50 ml are separate rows)",
        "• Same Group ID and Style ID for all sizes of one fragrance",
        "",
        "REVIEW",
        "• Update Inventory before upload",
        "• Confirm MRP matches your label",
        "• Shorten description if Meesho rejects long text",
    ]
    for i, line in enumerate(instructions, start=1):
        inst.cell(row=i, column=1, value=line)
    inst.column_dimensions["A"].width = 95

    # Catalog data
    catalog = wb.create_sheet("Catalog Data")
    style_header(catalog)
    rows = build_rows()
    headers = [c[0] for c in COLUMNS]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(headers, start=1):
            catalog.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
    autosize(catalog)
    catalog.freeze_panes = "A2"

    # Validation hints
    val = wb.create_sheet("Validation Hints")
    hints = [
        ("Field", "Guidance"),
        ("Variation / Capacity", "Use exact values from Meesho dropdown (e.g. 20 ml, 50 ml)"),
        ("HSN ID", "3303 = Perfumes and toilet waters (verify in Meesho dropdown)"),
        ("GST %", "18"),
        ("Meesho Price", "Your selling price on Meesho"),
        ("Wrong/Defective Returns Price", "~10% below Meesho Price; restricted returns"),
        ("MRP", "Must exceed Meesho Price; match label MRP incl. taxes"),
        ("Group ID", "Same for all sizes of one fragrance"),
        ("Style ID", "Same for all sizes; changes only for colour variants"),
        ("SKU ID", "Unique per size variant"),
        ("Image links", "From Meesho Image Bulk Upload only"),
        ("Product Description", "Keep short; no restricted keywords"),
        ("Country of Origin", "India"),
    ]
    for r, (a, b) in enumerate(hints, start=1):
        val.cell(row=r, column=1, value=a)
        val.cell(row=r, column=2, value=b)
    val.column_dimensions["A"].width = 34
    val.column_dimensions["B"].width = 70

    wb.save(OUT)
    print(f"Wrote {OUT} ({len(rows)} SKU rows)")


if __name__ == "__main__":
    main()
